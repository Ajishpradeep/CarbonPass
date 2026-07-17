#!/usr/bin/env python3
"""Caveat closure: prove the filled Communication Template recomputes correctly.

openpyxl writes formulas WITHOUT cached values, so any spreadsheet engine that
exports our filled workbook to CSV is forced to actually EVALUATE the template's
formula chain (including the hidden InputOutput array-formula matrix). We export
Summary_Products headlessly via LibreOffice and diff the SEE columns against the
engine's numbers in the .flags.json sidecar.

Pass = the Commission's workbook, filled by our writer, independently computes
the same SEE as our rules engine (which is itself golden-tested against the
Commission's worked example).

Usage:
    python scripts/verify_workbook_recalc.py [out/firm_a_communication_template.xlsx]
Exit 0 on match (rel tol 1e-6), 1 on mismatch/failure.
"""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

REL_TOL = 1e-6
SOFFICE_CANDIDATES = [
    "soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
]


def find_soffice() -> str | None:
    for cand in SOFFICE_CANDIDATES:
        if shutil.which(cand) or Path(cand).exists():
            return cand
    return None


def recalc_via_roundtrip(soffice: str, xlsx: Path, outdir: Path) -> Path:
    """xlsx -> (LibreOffice load: full recalculation, since openpyxl saved the
    formulas without cached values) -> xlsx with cached values at full precision."""
    cmd = [soffice, "--headless", "--norestore",
           "--convert-to", "xlsx:Calc MS Excel 2007 XML",
           "--outdir", str(outdir), str(xlsx)]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
    out = outdir / xlsx.name
    if not out.exists():
        raise RuntimeError(f"soffice convert failed: {r.stdout}\n{r.stderr}")
    return out


def main() -> int:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "out/firm_a_communication_template.xlsx")
    sidecar = xlsx.with_suffix(".flags.json")
    if not xlsx.exists() or not sidecar.exists():
        print(f"[recalc] missing {xlsx} or sidecar — run `carbonpass pack` first", file=sys.stderr)
        return 1
    soffice = find_soffice()
    if soffice is None:
        print("[recalc] LibreOffice (soffice) not found — install with "
              "`brew install --cask libreoffice`", file=sys.stderr)
        return 1

    expected = json.loads(sidecar.read_text(encoding="utf-8"))["products"]

    import openpyxl

    with tempfile.TemporaryDirectory() as td:
        recalced = recalc_via_roundtrip(soffice, xlsx, Path(td))
        wb = openpyxl.load_workbook(recalced, data_only=True)  # cached values = LO's recalc
        n_sheets = len(wb.sheetnames)
        ws = wb["Summary_Products"]

        print(f"[recalc] {xlsx.name}: {n_sheets} sheets; recalculated via LibreOffice round-trip")
        failures = 0
        for i, prod in enumerate(expected):
            r = 10 + i
            got = {
                "product": ws[f"D{r}"].value or "",
                "cn": ws[f"F{r}"].value or "",
                "see_direct": ws[f"I{r}"].value,
                "see_indirect": ws[f"J{r}"].value,
                "see_total": ws[f"K{r}"].value,
                "embedded_elec": ws[f"O{r}"].value,
            }
            want = {
                "see_direct": prod["see_direct"]["value"],
                "see_indirect": prod["see_indirect"]["value"],
                "see_total": prod["see_total"]["value"],
                "embedded_elec": prod["embedded_electricity_mwh_per_t"]["value"],
            }
            print(f"  row {r}: {got['product']!r} CN {got['cn']}")
            for k, w in want.items():
                g = got[k]
                ok = g is not None and (abs(g - w) <= REL_TOL * max(abs(w), 1e-12))
                status = "OK " if ok else "FAIL"
                print(f"    {status} {k:14} workbook={g!r:<24} engine={w!r}")
                failures += 0 if ok else 1

    if failures:
        print(f"[recalc] MISMATCH: {failures} value(s) differ — writer or engine bug", file=sys.stderr)
        return 1
    print("[recalc] PASS: workbook formulas independently reproduce the engine's SEE "
          f"(rel tol {REL_TOL}). Caveat closed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
