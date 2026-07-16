#!/usr/bin/env python3
"""Dump + classify the CBAM Communication Template workbooks.

For each sheet of interest, emits a JSON dump to out/template_dump/<sheet>.json with,
per non-empty cell of the EXAMPLE workbook:
    coord, value (example, computed), formula (if any), blank_value (blank template),
    kind: "formula" | "input" | "static"
        formula  -> cell holds a workbook formula (writer must NOT touch)
        input    -> literal in the example that differs from the blank template
                    (i.e. hand-entered by the operator — the cells our writer fills)
        static   -> identical literal in both (labels, headers, prefilled scaffolding)

Usage:  python scripts/inspect_template.py [SHEET ...]   (default: the 7 sheets we fill)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
BLANK = REPO / "data/cbam_official/communication_template.xlsx"
EXAMPLE = (
    REPO
    / "data/cbam_official/template_examples/4 CBAM SEE V2.1_Example Steel 3 Screws and nuts_final.xlsx"
)
OUT = REPO / "out/template_dump"

DEFAULT_SHEETS = [
    "A_InstData",
    "B_EmInst",
    "C_Emissions&Energy",
    "D_Processes",
    "E_PurchPrec",
    "Summary_Processes",
    "Summary_Products",
    "Summary_Communication",
]


def main() -> int:
    sheets = sys.argv[1:] or DEFAULT_SHEETS
    OUT.mkdir(parents=True, exist_ok=True)

    ex_vals = openpyxl.load_workbook(EXAMPLE, data_only=True)   # cached formula results
    ex_form = openpyxl.load_workbook(EXAMPLE, data_only=False)  # formulas
    bl_vals = openpyxl.load_workbook(BLANK, data_only=True)

    for name in sheets:
        ws_v, ws_f, ws_b = ex_vals[name], ex_form[name], bl_vals[name]
        cells = []
        for row in ws_v.iter_rows():
            for c in row:
                if c.value is None or str(c.value).strip() == "":
                    continue
                coord = c.coordinate
                f = ws_f[coord].value
                b = ws_b[coord].value
                is_formula = isinstance(f, str) and f.startswith("=")
                if is_formula:
                    kind = "formula"
                elif b is None or str(b) != str(c.value):
                    kind = "input"
                else:
                    kind = "static"
                rec = {"coord": coord, "value": c.value, "kind": kind}
                if is_formula:
                    rec["formula"] = f
                if kind == "input" and b is not None:
                    rec["blank_value"] = b
                cells.append(rec)
        out_file = OUT / f"{name.replace('&', '_')}.json"
        out_file.write_text(
            json.dumps(cells, ensure_ascii=False, indent=1, default=str), encoding="utf-8"
        )
        kinds = {}
        for r in cells:
            kinds[r["kind"]] = kinds.get(r["kind"], 0) + 1
        print(f"{name:24} {len(cells):5} non-empty  {kinds}  -> {out_file.relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
