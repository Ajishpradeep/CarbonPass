"""Writer: fill the blank EU Communication Template with activity + SEE results.

Contract = schema/cbam_template_map.yaml. The writer touches INPUT cells only;
all SEE outputs (Summary_Products I/J/K/O, InputOutput matrix) are workbook
formulas that Excel/LibreOffice recompute on open. openpyxl does not evaluate
formulas, so the authoritative numbers for screens/tests come from
carbonpass.rules — the workbook and the engine implement the same math (proven
by tests/golden against the Commission's worked example).

Every written value gets:
  * an Excel cell comment: source (actual/default) + relative uncertainty
  * a line in the sidecar flags JSON (machine-readable verifier checklist),
    including needs-attention items (default-value usage, >20% default share).
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

from carbonpass.config import BLANK_TEMPLATE, GRID_EF_KGCO2_PER_KWH
from carbonpass.pack import load_activity, run_allocation, run_rules

# Row anchors from schema/cbam_template_map.yaml (keep in sync)
D_PROC_HEADER = 11
D_PROC_STRIDE = 65
E_PREC_HEADER = 14
E_PREC_STRIDE = 44
GOODS_ROW0 = 62
PROC_ROW0 = 83
PREC_ROW0 = 102
BEMINST_ROW0 = 17
SUMPROD_ROW0 = 10


class Filler:
    def __init__(self, wb):
        self.wb = wb
        self.flags: list[dict] = []

    def put(self, sheet: str, coord: str, value, source: str = "actual",
            uncertainty_rel: float = 0.0, note: str = "", flag: str | None = None):
        ws = self.wb[sheet]
        cell = ws[coord]
        existing = cell.value
        if isinstance(existing, str) and existing.startswith("="):
            raise RuntimeError(
                f"refusing to overwrite formula at {sheet}!{coord} ({existing[:60]}) — "
                f"writer must only touch input cells (see cbam_template_map.yaml)")
        cell.value = value
        status = flag or ("default" if source == "default" else "actual")
        txt = f"CarbonPass: {status}"
        if uncertainty_rel:
            txt += f", ±{uncertainty_rel:.1%} (1σ)"
        if note:
            txt += f" — {note}"
        cell.comment = Comment(txt, "CarbonPass", height=60, width=260)
        self.flags.append({
            "cell": f"{sheet}!{coord}", "value": str(value), "source": source,
            "uncertainty_rel": uncertainty_rel, "flag": status, "note": note,
        })


def fill_template(activity: dict, out_path: str | Path) -> dict:
    alloc = run_allocation(activity)
    results = run_rules(activity, alloc)
    agg = activity["aggregated"]
    inst = activity["installation"]
    period = activity["period"]

    wb = openpyxl.load_workbook(BLANK_TEMPLATE, data_only=False)
    f = Filler(wb)

    # --- A_InstData -----------------------------------------------------------
    f.put("A_InstData", "I9", date.fromisoformat(period["start"]))
    f.put("A_InstData", "L9", date.fromisoformat(period["end"]))
    ident_cells = {
        "I19": inst.get("name_zh"), "I20": inst["name_en"], "I21": inst.get("street"),
        "I22": inst.get("economic_activity"), "I23": inst.get("post_code"),
        "I25": inst.get("city"), "I26": inst.get("country"), "I27": inst.get("unlocode"),
        "I28": inst.get("latitude"), "I29": inst.get("longitude"),
    }
    for coord, v in ident_cells.items():
        if v:
            f.put("A_InstData", coord, v, note="owner-provided (onboarding)")

    # G62 ("Route") is a workbook formula: with no route boxes ticked in I..N it
    # evaluates to "All production routes" — write the category only.
    f.put("A_InstData", f"E{GOODS_ROW0}", "Iron or steel products")

    production = agg["production"]
    for i, p in enumerate(production):
        r = PROC_ROW0 + i
        f.put("A_InstData", f"E{r}", "Iron or steel products")
        f.put("A_InstData", f"F{r}", "Only direct production")
        f.put("A_InstData", f"L{r}", p["process_name"])

    steel_inputs = agg.get("steel_inputs", [])
    for i, s in enumerate(steel_inputs):
        r = PREC_ROW0 + i
        f.put("A_InstData", f"E{r}", "Iron or steel products")
        f.put("A_InstData", f"F{r}", s.get("country", "TW"))
        f.put("A_InstData", f"L{r}", s["name"])

    # --- B_EmInst source streams -----------------------------------------------
    for i, fuel in enumerate(agg.get("fuels", [])):
        r = BEMINST_ROW0 + i
        amt = fuel["amount"]
        f.put("B_EmInst", f"D{r}", fuel.get("method", "Combustion"))
        f.put("B_EmInst", f"E{r}", fuel["name"])
        f.put("B_EmInst", f"F{r}", amt["value"], source=amt["source"],
              uncertainty_rel=amt.get("uncertainty_rel", 0.0),
              note=f"from {amt.get('n_documents', '?')} e-invoices (MIG XML)")
        f.put("B_EmInst", f"G{r}", fuel["unit"])
        f.put("B_EmInst", f"H{r}", fuel["ncv_gj_per_unit"])
        f.put("B_EmInst", f"J{r}", fuel["ef_tco2_per_tj"],
              note=f"EF source: {fuel.get('ef_source', '')}")
        f.put("B_EmInst", f"K{r}", "tCO2/TJ")

    # --- C_Emissions&Energy ------------------------------------------------------
    f.put("C_Emissions&Energy", "J16", 0, note="no on-site electricity generation")
    f.put("C_Emissions&Energy", "K16", 0, note="no non-CBAM production")
    total_indirect = sum(a.electricity_mwh for a in alloc.processes) * GRID_EF_KGCO2_PER_KWH
    elec = agg["electricity_mwh"]
    f.put("C_Emissions&Energy", "M26", round(total_indirect, 6),
          source=elec["source"], uncertainty_rel=elec.get("uncertainty_rel", 0.0),
          note="Σ process electricity × grid EF 0.474 (MOEA 2024); manual entry per template")
    f.put("C_Emissions&Energy", "H40",
          "Mostly measurements & international standard factors for e.g. the emission factor")
    f.put("C_Emissions&Energy", "H42", "Four eyes principle")

    # --- D_Processes -------------------------------------------------------------
    alloc_by_name = {a.name: a for a in alloc.processes}
    for i, p in enumerate(production):
        h = D_PROC_HEADER + i * D_PROC_STRIDE
        a = alloc_by_name[p["process_name"]]
        t = p["tonnes"]
        f.put("D_Processes", f"L{h + 5}", t["value"], source=t["source"],
              uncertainty_rel=t.get("uncertainty_rel", 0.0), note="production log")
        f.put("D_Processes", f"K{h + 39}", False)
        f.put("D_Processes", f"L{h + 39}", False)
        f.put("D_Processes", f"L{h + 43}", round(a.direct_emissions_tco2e, 6),
              uncertainty_rel=a.direct_unc_rel,
              note=f"allocation engine: fuel share {a.fuel_share:.1%} by production-tonnage prior "
                   f"(Monte-Carlo n={alloc.n_samples})")
        f.put("D_Processes", f"L{h + 54}", round(a.electricity_mwh, 6),
              uncertainty_rel=a.electricity_unc_rel,
              note=f"allocation engine: electricity share {a.elec_share:.1%} by machine kW×h prior")
        f.put("D_Processes", f"L{h + 55}", GRID_EF_KGCO2_PER_KWH,
              note="Taiwan grid EF 2024, MOEA Energy Administration")
        f.put("D_Processes", f"L{h + 56}", "D.4(b)")
        f.put("D_Processes", f"L{h + 60}", 0, note="no electricity exported")

    # --- E_PurchPrec ---------------------------------------------------------------
    proc_names = [p["process_name"] for p in production]
    results_by_name = {r.product_name: r for r in results}
    for i, s in enumerate(steel_inputs):
        h = E_PREC_HEADER + i * E_PREC_STRIDE
        m = s["mass_t"]
        f.put("E_PurchPrec", f"L{h + 3}", m["value"], source=m["source"],
              uncertainty_rel=m.get("uncertainty_rel", 0.0),
              note=f"Σ {m.get('n_documents', '?')} steel e-invoices (MIG XML)")
        consumption = (s.get("consumption_t") or {}).get("consumed_t", {})
        if not consumption and len(proc_names) == 1:
            consumption = {proc_names[0]: m["value"]}
        for j, pname in enumerate(proc_names):
            if pname in consumption:
                f.put("E_PurchPrec", f"L{h + 14 + j}", consumption[pname],
                      note="owner-confirmed consumption split")
        other = (s.get("consumption_t") or {}).get("consumed_other_t", 0.0)
        f.put("E_PurchPrec", f"L{h + 24}", other)

        epd = s.get("epd")
        if epd:
            f.put("E_PurchPrec", f"L{h + 35}", epd["see_direct"], source="actual",
                  uncertainty_rel=0.05, note=f"mill EPD: {epd.get('document', '')}")
            f.put("E_PurchPrec", f"M{h + 35}", "Measured")
            f.put("E_PurchPrec", f"L{h + 36}", epd.get("spec_electricity_mwh_per_t", 0.0),
                  source="actual", note="mill EPD")
            f.put("E_PurchPrec", f"M{h + 36}", "Measured")
            f.put("E_PurchPrec", f"L{h + 37}", epd.get("electricity_ef", 0.0), source="actual")
            f.put("E_PurchPrec", f"M{h + 37}", "Measured")
        else:
            # default value with the period's mark-up, taken from the SEE results
            # (rules.defaults) so writer and engine cannot diverge.
            line = None
            for r in results:
                for pl in r.precursor_lines:
                    if pl["precursor"] == s["name"]:
                        line = pl["see_direct_used"]
            assert line is not None, f"no SEE line for precursor {s['name']}"
            f.put("E_PurchPrec", f"L{h + 35}", line["value"], source="default",
                  flag="needs-verifier-attention", note=line["note"])
            f.put("E_PurchPrec", f"M{h + 35}", "Default value")
            f.put("E_PurchPrec", f"L{h + 36}", 0, source="default",
                  note="defaults carry no indirect for steel (N/A)")
            f.put("E_PurchPrec", f"M{h + 36}", "Default value")
            f.put("E_PurchPrec", f"L{h + 37}", 0, source="default")
            f.put("E_PurchPrec", f"M{h + 37}", "Default value")
            f.put("E_PurchPrec", f"K{h + 40}",   # merged K:M input; E holds the label
                  "No mill EPD available from the precursor supplier for this period; "
                  "CBAM default value applied with the period mark-up as the conservative fallback.",
                  source="default", flag="needs-verifier-attention")

    # --- Summary_Products ------------------------------------------------------------
    for i, p in enumerate(production):
        r = SUMPROD_ROW0 + i
        res = results_by_name[p["process_name"]]
        f.put("Summary_Products", f"D{r}", p["process_name"])
        f.put("Summary_Products", f"F{r}", p["cn_code"])
        f.put("Summary_Products", f"H{r}", p.get("invoice_name", ""))
        # engine's numbers land in the sidecar (workbook computes its own in I/J/K)
        f.flags.append({
            "cell": f"Summary_Products!I{r}..K{r} (computed by workbook)",
            "value": f"SEE dir {res.see_direct.value:.6f} ind {res.see_indirect.value:.6f} "
                     f"total {res.see_total.value:.6f} tCO2e/t (engine)",
            "source": res.see_total.source,
            "uncertainty_rel": res.see_total.uncertainty_rel,
            "flag": "engine-crosscheck",
            "note": "workbook formulas must reproduce these on open (same math, golden-tested)",
        })

    # --- save + integrity check ---------------------------------------------------
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    check = openpyxl.load_workbook(out_path, data_only=False)
    assert len(check.sheetnames) == 19, f"sheet count {len(check.sheetnames)} != 19"
    probe = check["Summary_Products"]["I10"].value
    assert isinstance(probe, str) and probe.startswith("="), "Summary_Products!I10 formula lost"

    needs_attention = [x for r in results for x in r.needs_attention]
    sidecar = {
        "output": str(out_path),
        "period": period,
        "installation": inst["name_en"],
        "products": [{
            "product_name": r.product_name, "cn_code": r.cn_code,
            "see_direct": r.see_direct.__dict__, "see_indirect": r.see_indirect.__dict__,
            "see_total": r.see_total.__dict__,
            "embedded_electricity_mwh_per_t": r.embedded_electricity_mwh_per_t.__dict__,
            "share_default_values": r.share_default_values,
            "precursors": r.precursor_lines,
        } for r in results],
        "allocation_notes": alloc.notes,
        "needs_attention": needs_attention,
        "cells_written": f.flags,
        "disclaimer": "This pack PREPARES verification. Verification itself is the act of an "
                      "accredited verifier; nothing in this file is certified.",
    }
    sidecar_path = out_path.with_suffix(".flags.json")
    sidecar_path.write_text(json.dumps(sidecar, ensure_ascii=False, indent=2), encoding="utf-8")
    return sidecar


def run_pack_cli(activity_json: str, output: str | None) -> int:
    activity = load_activity(activity_json)
    stem = Path(activity_json).stem.replace("_activity", "")
    out = Path(output) if output else Path("out") / f"{stem}_communication_template.xlsx"
    sidecar = fill_template(activity, out)
    print(f"OK: filled template -> {out}")
    print(f"    flags sidecar   -> {out.with_suffix('.flags.json')}")
    for p in sidecar["products"]:
        print(f"    CN {p['cn_code']}: SEE direct {p['see_direct']['value']:.4f} "
              f"indirect {p['see_indirect']['value']:.4f} total {p['see_total']['value']:.4f} tCO2e/t "
              f"(default share {p['share_default_values']:.0%})")
    for n in sidecar["needs_attention"]:
        print(f"    [needs-attention] {n}")
    return 0
