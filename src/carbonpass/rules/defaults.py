"""CBAM default values (IR 2025/2621) — parser for data/cbam_official/default_values.xlsx.

One sheet per country; columns:
    A CN code | B description | C direct | D indirect | E total
    F 2026 (+10%) | G 2027 (+20%) | H 2028+ (+30%) | I production route

Indirect for iron & steel goods is 'N/A' — consistent with G7: indirect emissions
are not part of the CN 7318 certificate obligation.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache

import openpyxl

from carbonpass.config import DEFAULT_VALUES_XLSX

# The "Other countries and territories" table (Annex I). Excel truncates sheet names at
# 31 chars, hence the clipped spelling — this is the literal sheet name in the workbook.
FALLBACK_SHEET = "_Other Countries and Territorie"


@dataclass(frozen=True)
class DefaultValue:
    cn_code: str            # normalized, no spaces (may be 4..8 digits)
    description: str
    direct: float | None
    indirect: float | None  # None == N/A (not in certificate for this sector)
    total: float | None
    y2026: float | None     # incl. +10% mark-up
    y2027: float | None     # incl. +20%
    y2028: float | None     # incl. +30% (2028 onward)
    route: str

    def for_year(self, year: int) -> float | None:
        """Marked-up default value applicable to a determination period.

        ⚠️ The workbook's marked-up column applies to the TOTAL (direct + indirect),
        not to direct alone (docs/15 §2.7 / §6 defect 1). For iron & steel this is
        harmless (indirect is N/A, so direct == total); for cement/fertiliser use
        `for_year_direct()` / `for_year_indirect()` for the split figures.
        """
        if year <= 2026:
            return self.y2026
        if year == 2027:
            return self.y2027
        return self.y2028

    def derived_markup(self, year: int) -> float | None:
        """Mark-up implied by the row's own marked-up column: for_year/total − 1.

        Never hard-code 10/20/30: iron & steel carries 10/20/30% but fertilisers
        carry a flat 1% in every year (docs/15 §6 defect 2). The row is the only
        safe source.
        """
        mk = self.for_year(year)
        base = self.total if self.total is not None else self.direct
        if mk is None or not base:
            return None
        return mk / base - 1.0

    def for_year_direct(self, year: int) -> float | None:
        """DIRECT default with the period's row-derived mark-up applied."""
        m = self.derived_markup(year)
        if m is None or self.direct is None:
            return None
        return self.direct * (1.0 + m)

    def for_year_indirect(self, year: int) -> float | None:
        """INDIRECT default with the period's row-derived mark-up (None where N/A)."""
        m = self.derived_markup(year)
        if m is None or self.indirect is None:
            return None
        return self.indirect * (1.0 + m)


def _num(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    return None


@lru_cache(maxsize=8)
def load_country(country: str = "Taiwan") -> list[DefaultValue]:
    wb = openpyxl.load_workbook(DEFAULT_VALUES_XLSX, data_only=True, read_only=True)
    if country not in wb.sheetnames:
        raise KeyError(f"no sheet {country!r} in default_values.xlsx")
    out = []
    for row in wb[country].iter_rows(min_row=3):
        code_raw = str(row[0].value or "").strip()
        if not re.fullmatch(r"[0-9 ]{4,12}", code_raw):
            continue  # section headers etc.
        out.append(DefaultValue(
            cn_code=code_raw.replace(" ", ""),
            description=str(row[1].value or "").strip(),
            direct=_num(row[2].value),
            indirect=_num(row[3].value),
            total=_num(row[4].value),
            y2026=_num(row[5].value),
            y2027=_num(row[6].value),
            y2028=_num(row[7].value),
            route=str(row[8].value or "").strip(),
        ))
    wb.close()
    return out


def lookup(cn_code: str, country: str = "Taiwan") -> DefaultValue | None:
    """Longest-prefix match: '73181542' -> row '731815' if no exact row exists.

    Guard (docs/15 §6 defect 10): a query SHORTER than the stored row used to return
    None silently — `lookup("7223")` missed the row "722300" and the caller skipped a
    precursor. Now, when the forward match fails, rows that EXTEND the query are
    considered: exactly one -> use it; several -> raise (ambiguous short code, pass
    more digits); none -> None, meaning genuinely unlisted, so `resolve()`'s country
    fallback stays legitimate (IR 2025/2621 Annex I preamble).
    """
    cn = cn_code.replace(" ", "")
    rows = load_country(country)
    best: DefaultValue | None = None
    for r in rows:
        if cn.startswith(r.cn_code) and r.direct is not None:
            if best is None or len(r.cn_code) > len(best.cn_code):
                best = r
    if best is not None:
        return best
    # reverse match: stored rows that extend a short query
    extending = [r for r in rows if r.cn_code.startswith(cn) and r.direct is not None]
    if len(extending) == 1:
        return extending[0]
    if len(extending) > 1:
        codes = sorted({r.cn_code for r in extending})
        raise ValueError(
            f"CN {cn_code!r} is ambiguous in the {country} sheet — {len(codes)} rows extend "
            f"it ({', '.join(codes[:6])}{'…' if len(codes) > 6 else ''}); pass more digits")
    return None


def resolve(cn_code: str, country: str = "Taiwan") -> tuple[DefaultValue | None, bool]:
    """Country value, else the 'Other countries and territories' table. -> (dv, used_fallback)

    Implements the rule in IR (EU) 2025/2621 Annex I preamble (verified in the legal text,
    docs/15 §8.1 Gate B; same rule in the Commission's Q&A p.37 §4.25):

        "Where a country or territory is explicitly listed but no value is provided or the
         relevant field shows '–', the default value for the respective good from the table
         'Other countries and territories' needs to be selected."

    That table is "the average of the ten exporting countries with the highest emission
    intensities per good" (Q&A p.37 §4.26) — i.e. reaching it is expensive.

    This is not hypothetical for Taiwan: of the 33 countries with a full steel book, only
    Taiwan, Thailand and Vietnam have **no CN 7221 value at all** (stainless wire rod —
    what a stainless fastener maker buys). Taiwan's 7221 row in the published OJ reads
    literally `see below ... #VALUE! #VALUE! #VALUE!` — a demonstrated publishing defect,
    in law. So every Taiwanese stainless fastener resolves here: 4.82 -> 5.302 in 2026.
    """
    dv = lookup(cn_code, country)
    if dv is not None:
        return dv, False
    dv = lookup(cn_code, FALLBACK_SHEET)
    return dv, dv is not None
